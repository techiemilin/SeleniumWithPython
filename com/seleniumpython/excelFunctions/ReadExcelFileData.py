'''
Created on Apr. 10, 2019

@author: milinpatel
''' 

import openpyxl 

path = "/Users/milinpatel/Documents/workspace/SeleniumWithPython/excelfiles/PythonDataDriven.xlsx"

workbook = openpyxl.load_workbook(path)

# sheet = workbook.active ***** for only one xcel sheet and active sheet

sheet = workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
columns = sheet.max_column

print(rows)
print(columns)

for r in range (1, rows + 1):
    for c in range (1, columns + 1):
        print(sheet.cell(r, c).value, end="     ")
        
    print()
